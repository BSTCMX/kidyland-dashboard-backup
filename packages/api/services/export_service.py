"""
Export Service - Business logic for generating Excel and PDF reports.

Features:
- Excel export with multiple sheets (Ventas, Inventario, Servicios, Horas Pico, Productos Top, Servicios Top, Clientes Top)
- PDF export with all metrics in organized sections
- Modular design following Clean Architecture
- Reuses ReportService for data retrieval
"""
import logging
from datetime import date, datetime
from typing import Dict, Any, Optional, List
from io import BytesIO
from sqlalchemy.ext.asyncio import AsyncSession

# Excel generation
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# PDF generation
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

from services.report_service import ReportService

logger = logging.getLogger(__name__)


class ExportService:
    """Service for generating Excel and PDF exports."""
    
    def __init__(self):
        """Initialize ExportService with ReportService."""
        self.report_service = ReportService()
    
    async def generate_excel_report(
        self,
        db: AsyncSession,
        sucursal_id: Optional[str] = None,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        report_type: str = "dashboard",
        sections: Optional[List[str]] = None,
        include_predictions: bool = False,
        module: Optional[str] = None
    ) -> BytesIO:
        """
        Generate Excel report with all metrics.
        
        Args:
            db: Database session
            sucursal_id: Optional sucursal ID to filter by
            start_date: Optional start date
            end_date: Optional end date
            report_type: Type of report (dashboard, sales, stock, services, summary, arqueos, customers, forecasting)
            sections: Optional list of specific sections to include. If None, exports all sections for report_type.
            include_predictions: Whether to include predictions if available
            module: Optional module filter ("recepcion", "kidibar", "all")
            
        Returns:
            BytesIO buffer with Excel file
        """
        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Determine sections to export (backward compatible)
        if sections is None:
            sections = self._get_default_sections(report_type, include_predictions)
        
        # Process each section requested
        await self._process_sections_for_excel(
            wb=wb,
            sections=sections,
            db=db,
            sucursal_id=sucursal_id,
            start_date=start_date,
            end_date=end_date,
            report_type=report_type,
            module=module
        )
        
        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        logger.info(f"Excel report generated: {report_type}, sections: {sections}")
        return buffer
    
    async def generate_pdf_report(
        self,
        db: AsyncSession,
        sucursal_id: Optional[str] = None,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        report_type: str = "dashboard",
        sections: Optional[List[str]] = None,
        include_predictions: bool = False,
        module: Optional[str] = None
    ) -> BytesIO:
        """
        Generate PDF report with all metrics.
        
        Args:
            db: Database session
            sucursal_id: Optional sucursal ID to filter by
            start_date: Optional start date
            end_date: Optional end date
            report_type: Type of report (dashboard, sales, stock, services)
            
        Returns:
            BytesIO buffer with PDF file
        """
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        story = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#1a1a2e'),
            spaceAfter=12,
            alignment=TA_CENTER
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#16213e'),
            spaceAfter=8,
            spaceBefore=12
        )
        
        # Title
        story.append(Paragraph("Reporte Kidyland", title_style))
        story.append(Spacer(1, 0.2 * inch))
        
        # Determine sections to export (backward compatible)
        if sections is None:
            sections = self._get_default_sections(report_type, include_predictions)
        
        # Process each section requested
        await self._process_sections_for_pdf(
            story=story,
            sections=sections,
            db=db,
            sucursal_id=sucursal_id,
            start_date=start_date,
            end_date=end_date,
            report_type=report_type,
            heading_style=heading_style,
            module=module,
            include_predictions=include_predictions
        )
        
        # Footer
        story.append(Spacer(1, 0.3 * inch))
        story.append(Paragraph(
            f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}",
            styles['Normal']
        ))
        
        # Build PDF
        doc.build(story)
        buffer.seek(0)
        
        logger.info(f"PDF report generated: {report_type}, sections: {sections}")
        return buffer
    
    # ========== HELPER METHODS ==========
    
    async def _fetch_forecasting_predictions(
        self,
        db: AsyncSession,
        sucursal_id: Optional[str],
        start_date: Optional[date],
        end_date: Optional[date],
        module: Optional[str],
        forecast_days: int = 7
    ) -> Dict[str, Any]:
        """
        Fetch or generate segmented forecasting predictions for export.
        
        Uses PredictionService to generate predictions on-demand.
        Handles multiple modules (recepcion, kidibar, total) and prediction types (sales, capacity, stock).
        
        Args:
            db: Database session
            sucursal_id: Optional sucursal ID to filter by
            start_date: Optional start date (not used directly but kept for consistency)
            end_date: Optional end date (not used directly but kept for consistency)
            module: Optional module filter ("recepcion", "kidibar", "all", or None for all)
            forecast_days: Number of days to forecast (default: 7)
            
        Returns:
            Dictionary with segmented predictions structure:
            {
                "predictions": {
                    "recepcion": {"sales": {...}, "capacity": {...}},
                    "kidibar": {"sales": {...}, "stock": {...}},
                    "total": {"sales": {...}}
                },
                "forecast_days": int,
                "modules": List[str]
            }
        """
        from services.prediction_service import PredictionService
        prediction_service = PredictionService()
        
        # Determine modules to forecast
        if not module or module == "all":
            modules_to_forecast = ["recepcion", "kidibar", "total"]
        elif module == "recepcion":
            modules_to_forecast = ["recepcion"]
        elif module == "kidibar":
            modules_to_forecast = ["kidibar"]
        else:
            modules_to_forecast = [module]
        
        # Generate predictions for each module
        segmented_predictions: Dict[str, Dict[str, Any]] = {}
        
        for mod in modules_to_forecast:
            module_predictions: Dict[str, Any] = {}
            
            # Generate sales predictions for each module
            try:
                sales_pred = await prediction_service.predict_sales_enhanced(
                    db=db,
                    sucursal_id=sucursal_id,
                    forecast_days=forecast_days,
                    module=mod if mod != "total" else None  # "total" means all modules aggregated (None)
                )
                module_predictions["sales"] = sales_pred
            except Exception as e:
                logger.warning(f"Could not generate sales predictions for {mod}: {e}", exc_info=True)
                module_predictions["sales"] = {
                    "error": "generation_failed",
                    "message": f"No se pudieron generar predicciones de ventas para {mod}",
                    "forecast": [],
                    "confidence": "low"
                }
            
            # Generate capacity predictions (only for recepcion or total)
            if mod in ["recepcion", "total"]:
                try:
                    capacity_pred = await prediction_service.predict_capacity(
                        db=db,
                        sucursal_id=sucursal_id,
                        forecast_days=forecast_days
                    )
                    module_predictions["capacity"] = capacity_pred
                except Exception as e:
                    logger.warning(f"Could not generate capacity predictions for {mod}: {e}", exc_info=True)
                    module_predictions["capacity"] = {
                        "error": "generation_failed",
                        "message": f"No se pudieron generar predicciones de capacidad para {mod}",
                        "forecast": [],
                        "confidence": "low"
                    }
            
            # Generate stock predictions (only for kidibar or total)
            if mod in ["kidibar", "total"]:
                try:
                    stock_pred = await prediction_service.predict_stock_needs(
                        db=db,
                        sucursal_id=sucursal_id,
                        forecast_days=forecast_days
                    )
                    module_predictions["stock"] = stock_pred
                except Exception as e:
                    logger.warning(f"Could not generate stock predictions for {mod}: {e}", exc_info=True)
                    module_predictions["stock"] = {
                        "error": "generation_failed",
                        "message": f"No se pudieron generar predicciones de stock para {mod}",
                        "reorder_suggestions": [],
                        "confidence": "low"
                    }
            
            segmented_predictions[mod] = module_predictions
        
        return {
            "predictions": segmented_predictions,
            "forecast_days": forecast_days,
            "modules": modules_to_forecast
        }
    
    async def _process_sections_for_excel(
        self,
        wb: Workbook,
        sections: List[str],
        db: AsyncSession,
        sucursal_id: Optional[str],
        start_date: Optional[date],
        end_date: Optional[date],
        report_type: str,
        module: Optional[str] = None
    ) -> None:
        """
        Process requested sections and create Excel sheets.
        
        This method handles fetching data and creating sheets for each requested section.
        """
        # Dashboard sections
        if "sales" in sections:
            sales_report = await self.report_service.get_sales_report(
                db=db,
                sucursal_id=sucursal_id,
                start_date=start_date,
                end_date=end_date,
                use_cache=False
            )
            self._create_sales_sheet(wb, sales_report)
        
        if "stock" in sections or "inventory" in sections:
            stock_report = await self.report_service.get_stock_report(
                db=db,
                sucursal_id=sucursal_id,
                use_cache=False
            )
            self._create_stock_sheet(wb, stock_report)
        
        if "services" in sections:
            services_report = await self.report_service.get_services_report(
                db=db,
                sucursal_id=sucursal_id,
                use_cache=False
            )
            self._create_services_sheet(wb, services_report)
        
        if "peak_hours" in sections:
            peak_hours_report = await self.report_service.get_peak_hours_report(
                db=db,
                sucursal_id=sucursal_id,
                target_date=end_date or date.today(),
                use_cache=False
            )
            self._create_peak_hours_sheet(wb, peak_hours_report)
        
        if "top_products" in sections:
            top_products_report = await self.report_service.get_top_products_report(
                db=db,
                sucursal_id=sucursal_id,
                days=7,
                use_cache=False
            )
            self._create_top_products_sheet(wb, top_products_report)
        
        if "top_services" in sections:
            top_services_report = await self.report_service.get_top_services_report(
                db=db,
                sucursal_id=sucursal_id,
                days=7,
                use_cache=False
            )
            self._create_top_services_sheet(wb, top_services_report)
        
        if "top_customers" in sections:
            top_customers_report = await self.report_service.get_top_customers_report(
                db=db,
                sucursal_id=sucursal_id,
                days=7,
                use_cache=False
            )
            self._create_top_customers_sheet(wb, top_customers_report)
        
        # Phase 5: Additional sections
        if "summary" in sections:
            summary_report = await self.report_service.get_dashboard_summary(
                db=db,
                sucursal_id=sucursal_id,
                use_cache=False
            )
            self._create_executive_summary_sheet(wb, summary_report)
        
        if "customers" in sections:
            customers_summary = await self.report_service.get_customers_summary(
                db=db,
                sucursal_id=sucursal_id,
                start_date=start_date,
                end_date=end_date,
                use_cache=False
            )
            self._create_customers_summary_sheet(wb, customers_summary)
        
        if "arqueos" in sections:
            arqueos_report = await self.report_service.get_arqueos_report(
                db=db,
                sucursal_id=sucursal_id,
                start_date=start_date,
                end_date=end_date,
                module=module,
                use_cache=False
            )
            self._create_arqueos_sheet(wb, arqueos_report)
        
        if "forecasting" in sections:
            try:
                forecasting_data = await self._fetch_forecasting_predictions(
                    db=db,
                    sucursal_id=sucursal_id,
                    start_date=start_date,
                    end_date=end_date,
                    module=module,
                    forecast_days=7
                )
                self._create_forecasting_sheet(wb, forecasting_data)
            except Exception as e:
                logger.error(f"Could not add forecasting section to Excel: {e}", exc_info=True)
                # Continue without forecasting if there's an error
    
    async def _process_sections_for_pdf(
        self,
        story: list,
        sections: List[str],
        db: AsyncSession,
        sucursal_id: Optional[str],
        start_date: Optional[date],
        end_date: Optional[date],
        report_type: str,
        heading_style: Any,
        module: Optional[str] = None,
        include_predictions: bool = False
    ) -> None:
        """
        Process requested sections and add them to PDF story.
        
        This method handles fetching data and adding PDF sections for each requested section.
        Similar to _process_sections_for_excel but for PDF format.
        """
        # Dashboard sections
        if "sales" in sections:
            sales_report = await self.report_service.get_sales_report(
                db=db,
                sucursal_id=sucursal_id,
                start_date=start_date,
                end_date=end_date,
                use_cache=False
            )
            self._add_sales_section_pdf(story, sales_report, heading_style)
        
        if "stock" in sections or "inventory" in sections:
            stock_report = await self.report_service.get_stock_report(
                db=db,
                sucursal_id=sucursal_id,
                use_cache=False
            )
            self._add_stock_section_pdf(story, stock_report, heading_style)
        
        if "services" in sections:
            services_report = await self.report_service.get_services_report(
                db=db,
                sucursal_id=sucursal_id,
                use_cache=False
            )
            self._add_services_section_pdf(story, services_report, heading_style)
        
        if "peak_hours" in sections:
            peak_hours_report = await self.report_service.get_peak_hours_report(
                db=db,
                sucursal_id=sucursal_id,
                target_date=end_date or date.today(),
                use_cache=False
            )
            self._add_peak_hours_section_pdf(story, peak_hours_report, heading_style)
        
        if "top_products" in sections:
            top_products_report = await self.report_service.get_top_products_report(
                db=db,
                sucursal_id=sucursal_id,
                days=7,
                use_cache=False
            )
            self._add_top_products_section_pdf(story, top_products_report, heading_style)
        
        if "top_services" in sections:
            top_services_report = await self.report_service.get_top_services_report(
                db=db,
                sucursal_id=sucursal_id,
                days=7,
                use_cache=False
            )
            self._add_top_services_section_pdf(story, top_services_report, heading_style)
        
        if "top_customers" in sections:
            top_customers_report = await self.report_service.get_top_customers_report(
                db=db,
                sucursal_id=sucursal_id,
                days=7,
                use_cache=False
            )
            self._add_top_customers_section_pdf(story, top_customers_report, heading_style)
        
        # Predictions section (if available and requested)
        if "predictions" in sections and include_predictions:
            try:
                from services.prediction_service import PredictionService
                prediction_service = PredictionService()
                
                # Fetch dashboard predictions
                predictions = await prediction_service.generate_all_predictions(
                    db=db,
                    sucursal_id=sucursal_id,
                    forecast_days=7
                )
                
                if predictions:
                    self._add_predictions_section_pdf(story, predictions, heading_style)
            except Exception as e:
                logger.warning(f"Could not add predictions section to PDF: {e}", exc_info=True)
                # Continue without predictions if there's an error
        
        # Phase 5: Additional sections
        if "summary" in sections:
            summary_report = await self.report_service.get_dashboard_summary(
                db=db,
                sucursal_id=sucursal_id,
                use_cache=False
            )
            self._add_executive_summary_section_pdf(story, summary_report, heading_style)
        
        if "customers" in sections:
            customers_summary = await self.report_service.get_customers_summary(
                db=db,
                sucursal_id=sucursal_id,
                start_date=start_date,
                end_date=end_date,
                use_cache=False
            )
            self._add_customers_summary_section_pdf(story, customers_summary, heading_style)
        
        if "arqueos" in sections:
            arqueos_report = await self.report_service.get_arqueos_report(
                db=db,
                sucursal_id=sucursal_id,
                start_date=start_date,
                end_date=end_date,
                module=module,
                use_cache=False
            )
            self._add_arqueos_section_pdf(story, arqueos_report, heading_style)
        
        if "forecasting" in sections:
            try:
                forecasting_data = await self._fetch_forecasting_predictions(
                    db=db,
                    sucursal_id=sucursal_id,
                    start_date=start_date,
                    end_date=end_date,
                    module=module,
                    forecast_days=7
                )
                self._add_forecasting_section_pdf(story, forecasting_data, heading_style)
            except Exception as e:
                logger.error(f"Could not add forecasting section to PDF: {e}", exc_info=True)
                # Continue without forecasting if there's an error
    
    def _get_default_sections(
        self, 
        report_type: str, 
        include_predictions: bool = False
    ) -> List[str]:
        """
        Get default sections to export based on report type.
        
        Args:
            report_type: Type of report
            include_predictions: Whether to include predictions by default
            
        Returns:
            List of section IDs to export
        """
        if report_type == "dashboard":
            sections = [
                "sales", "stock", "services",
                "peak_hours", "top_products", "top_services", "top_customers"
            ]
            if include_predictions:
                sections.append("predictions")
            return sections
        elif report_type == "sales":
            return ["sales"]
        elif report_type == "stock":
            return ["stock"]
        elif report_type == "services":
            return ["services"]
        elif report_type == "summary":
            return ["summary"]
        elif report_type == "arqueos":
            return ["arqueos"]
        elif report_type == "customers":
            return ["customers"]
        elif report_type == "forecasting":
            return ["forecasting"]
        elif report_type == "reports":
            # All report sections
            sections = ["summary", "sales", "inventory", "services", "arqueos", "customers", "forecasting"]
            if include_predictions:
                sections.append("predictions")
            return sections
        else:
            # Default: dashboard sections
            return ["sales", "stock", "services"]
    
    # ========== EXCEL SHEET CREATION METHODS ==========
    
    def _create_sales_sheet(self, wb: Workbook, sales_report: Dict[str, Any]) -> None:
        """Create sales metrics sheet in Excel."""
        ws = wb.create_sheet("Ventas")
        
        # Header style
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Title
        ws['A1'] = "Reporte de Ventas"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:D1')
        
        row = 3
        
        if sales_report:
            # Summary
            ws[f'A{row}'] = "Total Revenue:"
            ws[f'B{row}'] = f"${(sales_report.get('total_revenue_cents', 0) / 100):.2f}"
            row += 1
            
            ws[f'A{row}'] = "Total Ventas:"
            ws[f'B{row}'] = sales_report.get('sales_count', 0)
            row += 1
            
            ws[f'A{row}'] = "Ticket Promedio:"
            ws[f'B{row}'] = f"${(sales_report.get('average_transaction_value_cents', 0) / 100):.2f}"
            row += 2
            
            # Revenue by type
            if sales_report.get('revenue_by_type'):
                ws[f'A{row}'] = "Revenue por Tipo"
                ws[f'A{row}'].font = Font(bold=True)
                row += 1
                
                ws[f'A{row}'] = "Tipo"
                ws[f'B{row}'] = "Revenue"
                for cell in ws[f'A{row}:B{row}']:
                    for c in cell:
                        c.fill = header_fill
                        c.font = header_font
                row += 1
                
                for tipo, revenue in sales_report['revenue_by_type'].items():
                    ws[f'A{row}'] = tipo
                    ws[f'B{row}'] = f"${(revenue / 100):.2f}"
                    row += 1
        
        # Auto-adjust columns
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_stock_sheet(self, wb: Workbook, stock_report: Dict[str, Any]) -> None:
        """Create stock metrics sheet in Excel."""
        ws = wb.create_sheet("Inventario")
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        ws['A1'] = "Reporte de Inventario"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:D1')
        
        row = 3
        
        if stock_report:
            ws[f'A{row}'] = "Total Productos:"
            ws[f'B{row}'] = stock_report.get('total_products', 0)
            row += 1
            
            ws[f'A{row}'] = "Valor Total:"
            ws[f'B{row}'] = f"${(stock_report.get('total_stock_value_cents', 0) / 100):.2f}"
            row += 1
            
            ws[f'A{row}'] = "Alertas:"
            ws[f'B{row}'] = stock_report.get('alerts_count', 0)
            row += 2
            
            # Low stock alerts
            if stock_report.get('low_stock_alerts'):
                ws[f'A{row}'] = "Productos con Stock Bajo"
                ws[f'A{row}'].font = Font(bold=True)
                row += 1
                
                ws[f'A{row}'] = "Producto"
                ws[f'B{row}'] = "Stock Actual"
                ws[f'C{row}'] = "Umbral"
                for cell in ws[f'A{row}:C{row}']:
                    for c in cell:
                        c.fill = header_fill
                        c.font = header_font
                row += 1
                
                for alert in stock_report['low_stock_alerts']:
                    ws[f'A{row}'] = alert.get('product_name', '')
                    ws[f'B{row}'] = alert.get('stock_qty', 0)
                    ws[f'C{row}'] = alert.get('threshold_alert_qty', 0)
                    row += 1
        
        # Auto-adjust columns
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_services_sheet(self, wb: Workbook, services_report: Dict[str, Any]) -> None:
        """Create services metrics sheet in Excel."""
        ws = wb.create_sheet("Servicios")
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        ws['A1'] = "Reporte de Servicios"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:D1')
        
        row = 3
        
        if services_report:
            ws[f'A{row}'] = "Timers Activos:"
            ws[f'B{row}'] = services_report.get('active_timers_count', 0)
            row += 1
            
            ws[f'A{row}'] = "Total Servicios:"
            ws[f'B{row}'] = services_report.get('total_services', 0)
            row += 2
            
            # Services by sucursal
            if services_report.get('services_by_sucursal'):
                ws[f'A{row}'] = "Servicios por Sucursal"
                ws[f'A{row}'].font = Font(bold=True)
                row += 1
                
                ws[f'A{row}'] = "Sucursal"
                ws[f'B{row}'] = "Cantidad"
                for cell in ws[f'A{row}:B{row}']:
                    for c in cell:
                        c.fill = header_fill
                        c.font = header_font
                row += 1
                
                for sucursal, count in services_report['services_by_sucursal'].items():
                    ws[f'A{row}'] = sucursal
                    ws[f'B{row}'] = count
                    row += 1
        
        # Auto-adjust columns
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_peak_hours_sheet(self, wb: Workbook, peak_hours_report: Dict[str, Any]) -> None:
        """Create peak hours metrics sheet in Excel."""
        ws = wb.create_sheet("Horas Pico")
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        ws['A1'] = "Reporte de Horas Pico"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:D1')
        
        row = 3
        
        if peak_hours_report:
            busiest_hour = peak_hours_report.get('busiest_hour', {})
            ws[f'A{row}'] = "Hora Más Ocupada:"
            ws[f'B{row}'] = f"{busiest_hour.get('hour', 0)}:00h"
            row += 1
            
            ws[f'A{row}'] = "Ventas en Hora Pico:"
            ws[f'B{row}'] = busiest_hour.get('sales_count', 0)
            row += 2
            
            # Top peak hours
            if peak_hours_report.get('peak_hours'):
                ws[f'A{row}'] = "Top 5 Horas"
                ws[f'A{row}'].font = Font(bold=True)
                row += 1
                
                ws[f'A{row}'] = "Hora"
                ws[f'B{row}'] = "Ventas"
                ws[f'C{row}'] = "Revenue"
                for cell in ws[f'A{row}:C{row}']:
                    for c in cell:
                        c.fill = header_fill
                        c.font = header_font
                row += 1
                
                for peak in peak_hours_report['peak_hours']:
                    ws[f'A{row}'] = f"{peak.get('hour', 0)}:00h"
                    ws[f'B{row}'] = peak.get('sales_count', 0)
                    ws[f'C{row}'] = f"${(peak.get('revenue_cents', 0) / 100):.2f}"
                    row += 1
        
        # Auto-adjust columns
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_top_products_sheet(self, wb: Workbook, top_products_report: Dict[str, Any]) -> None:
        """Create top products metrics sheet in Excel."""
        ws = wb.create_sheet("Productos Top")
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        ws['A1'] = f"Productos Top - Últimos {top_products_report.get('period_days', 7)} días"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:D1')
        
        row = 3
        
        if top_products_report and top_products_report.get('top_products'):
            ws[f'A{row}'] = "Rank"
            ws[f'B{row}'] = "Producto"
            ws[f'C{row}'] = "Cantidad Vendida"
            ws[f'D{row}'] = "Revenue"
            for cell in ws[f'A{row}:D{row}']:
                for c in cell:
                    c.fill = header_fill
                    c.font = header_font
            row += 1
            
            for idx, product in enumerate(top_products_report['top_products'], 1):
                ws[f'A{row}'] = idx
                ws[f'B{row}'] = product.get('product_name', '')
                ws[f'C{row}'] = product.get('quantity_sold', 0)
                ws[f'D{row}'] = f"${(product.get('revenue_cents', 0) / 100):.2f}"
                row += 1
        
        # Auto-adjust columns
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_top_services_sheet(self, wb: Workbook, top_services_report: Dict[str, Any]) -> None:
        """Create top services metrics sheet in Excel."""
        ws = wb.create_sheet("Servicios Top")
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        ws['A1'] = f"Servicios Top - Últimos {top_services_report.get('period_days', 7)} días"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:D1')
        
        row = 3
        
        if top_services_report and top_services_report.get('top_services'):
            ws[f'A{row}'] = "Rank"
            ws[f'B{row}'] = "Servicio"
            ws[f'C{row}'] = "Usos"
            ws[f'D{row}'] = "Duración Promedio (min)"
            for cell in ws[f'A{row}:D{row}']:
                for c in cell:
                    c.fill = header_fill
                    c.font = header_font
            row += 1
            
            for idx, service in enumerate(top_services_report['top_services'], 1):
                ws[f'A{row}'] = idx
                ws[f'B{row}'] = service.get('service_name', '')
                ws[f'C{row}'] = service.get('usage_count', 0)
                ws[f'D{row}'] = f"{service.get('avg_duration_minutes', 0):.1f}"
                row += 1
        
        # Auto-adjust columns
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_top_customers_sheet(self, wb: Workbook, top_customers_report: Dict[str, Any]) -> None:
        """Create top customers metrics sheet in Excel."""
        ws = wb.create_sheet("Clientes Top")
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        ws['A1'] = f"Clientes Top - Últimos {top_customers_report.get('period_days', 7)} días"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:E1')
        
        row = 3
        
        if top_customers_report and top_customers_report.get('top_customers'):
            ws[f'A{row}'] = "Rank"
            ws[f'B{row}'] = "Nombre"
            ws[f'C{row}'] = "Edad"
            ws[f'D{row}'] = "Visitas"
            ws[f'E{row}'] = "Total Gastado"
            for cell in ws[f'A{row}:E{row}']:
                for c in cell:
                    c.fill = header_fill
                    c.font = header_font
            row += 1
            
            for idx, customer in enumerate(top_customers_report['top_customers'], 1):
                ws[f'A{row}'] = idx
                ws[f'B{row}'] = customer.get('child_name', '')
                ws[f'C{row}'] = customer.get('child_age', '') if customer.get('child_age') else 'N/A'
                ws[f'D{row}'] = customer.get('visit_count', 0)
                ws[f'E{row}'] = f"${((customer.get('total_revenue_cents', 0) or 0) / 100):.2f}"
                row += 1
        
        # Auto-adjust columns
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_executive_summary_sheet(self, wb: Workbook, summary_report: Dict[str, Any]) -> None:
        """Create executive summary sheet in Excel."""
        ws = wb.create_sheet("Resumen Ejecutivo")
        
        # Header style
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Title
        ws['A1'] = "Resumen Ejecutivo"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:D1')
        
        row = 3
        
        if summary_report:
            # Format generated_at date
            generated_at_str = "N/A"
            if summary_report.get('generated_at'):
                try:
                    from datetime import datetime
                    dt = datetime.fromisoformat(summary_report['generated_at'].replace('Z', '+00:00'))
                    generated_at_str = dt.strftime('%d/%m/%Y %H:%M:%S')
                except:
                    generated_at_str = str(summary_report.get('generated_at', 'N/A'))
            
            ws[f'A{row}'] = "Fecha de Generación:"
            ws[f'B{row}'] = generated_at_str
            row += 2
            
            # Sales section
            sales = summary_report.get('sales')
            if sales:
                ws[f'A{row}'] = "VENTAS"
                ws[f'A{row}'].font = Font(bold=True, size=12)
                row += 1
                
                ws[f'A{row}'] = "Total Revenue:"
                ws[f'B{row}'] = f"${(sales.get('total_revenue_cents', 0) / 100):.2f}"
                row += 1
                
                ws[f'A{row}'] = "Total Ventas:"
                ws[f'B{row}'] = sales.get('sales_count', 0)
                row += 1
                
                ws[f'A{row}'] = "Ticket Promedio:"
                ws[f'B{row}'] = f"${(sales.get('average_transaction_value_cents', 0) / 100):.2f}"
                row += 2
            
            # Stock section
            stock = summary_report.get('stock')
            if stock:
                ws[f'A{row}'] = "INVENTARIO"
                ws[f'A{row}'].font = Font(bold=True, size=12)
                row += 1
                
                ws[f'A{row}'] = "Total Productos:"
                ws[f'B{row}'] = stock.get('total_products', 0)
                row += 1
                
                ws[f'A{row}'] = "Valor Total:"
                ws[f'B{row}'] = f"${(stock.get('total_stock_value_cents', 0) / 100):.2f}"
                row += 1
                
                ws[f'A{row}'] = "Alertas:"
                ws[f'B{row}'] = stock.get('alerts_count', 0)
                row += 2
            
            # Services section
            services = summary_report.get('services')
            if services:
                ws[f'A{row}'] = "SERVICIOS"
                ws[f'A{row}'].font = Font(bold=True, size=12)
                row += 1
                
                ws[f'A{row}'] = "Timers Activos:"
                ws[f'B{row}'] = services.get('active_timers_count', 0)
                row += 1
                
                ws[f'A{row}'] = "Total Servicios:"
                ws[f'B{row}'] = services.get('total_services', 0)
                row += 1
        
        # Auto-adjust columns
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_customers_summary_sheet(self, wb: Workbook, customers_summary: Dict[str, Any]) -> None:
        """Create customers summary sheet in Excel."""
        ws = wb.create_sheet("Resumen Clientes")
        
        # Header style
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Title
        ws['A1'] = "Resumen de Clientes"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:D1')
        
        row = 3
        
        if customers_summary:
            # Main metrics
            ws[f'A{row}'] = "Total Clientes Únicos:"
            ws[f'B{row}'] = customers_summary.get('total_unique_customers', 0)
            row += 1
            
            ws[f'A{row}'] = "Nuevos Clientes:"
            ws[f'B{row}'] = customers_summary.get('new_customers', 0)
            row += 1
            
            avg_revenue = customers_summary.get('avg_revenue_per_customer_cents', 0) or 0
            ws[f'A{row}'] = "Revenue Promedio por Cliente:"
            ws[f'B{row}'] = f"${(avg_revenue / 100):.2f}"
            row += 1
            
            total_revenue = customers_summary.get('total_revenue_cents', 0) or 0
            ws[f'A{row}'] = "Revenue Total:"
            ws[f'B{row}'] = f"${(total_revenue / 100):.2f}"
            row += 2
            
            # Breakdown by module
            recepcion_count = customers_summary.get('recepcion_customers', 0) or 0
            kidibar_count = customers_summary.get('kidibar_customers', 0) or 0
            
            if recepcion_count > 0 or kidibar_count > 0:
                ws[f'A{row}'] = "DESGLOSE POR MÓDULO"
                ws[f'A{row}'].font = Font(bold=True, size=12)
                row += 1
                
                ws[f'A{row}'] = "Módulo"
                ws[f'B{row}'] = "Clientes"
                for cell in ws[f'A{row}:B{row}']:
                    for c in cell:
                        c.fill = header_fill
                        c.font = header_font
                row += 1
                
                ws[f'A{row}'] = "Recepción"
                ws[f'B{row}'] = recepcion_count
                row += 1
                
                ws[f'A{row}'] = "KidiBar"
                ws[f'B{row}'] = kidibar_count
                row += 1
        
        # Auto-adjust columns
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_arqueos_sheet(self, wb: Workbook, arqueos_report: Dict[str, Any]) -> None:
        """Create arqueos (day close) report sheet in Excel."""
        ws = wb.create_sheet("Arqueos")
        
        # Header style
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Title
        ws['A1'] = "Reporte de Arqueos"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:E1')
        
        row = 3
        
        if arqueos_report:
            # Period info
            period = arqueos_report.get('period', {})
            if period:
                start_date_str = period.get('start_date', 'N/A')
                end_date_str = period.get('end_date', 'N/A')
                ws[f'A{row}'] = "Período:"
                ws[f'B{row}'] = f"{start_date_str} - {end_date_str}"
                row += 2
            
            # Summary metrics
            ws[f'A{row}'] = "RESUMEN"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1
            
            ws[f'A{row}'] = "Total Arqueos:"
            ws[f'B{row}'] = arqueos_report.get('total_arqueos', 0)
            row += 1
            
            ws[f'A{row}'] = "Total Sistema:"
            ws[f'B{row}'] = f"${(arqueos_report.get('total_system_cents', 0) / 100):.2f}"
            row += 1
            
            ws[f'A{row}'] = "Total Físico:"
            ws[f'B{row}'] = f"${(arqueos_report.get('total_physical_cents', 0) / 100):.2f}"
            row += 1
            
            ws[f'A{row}'] = "Diferencia Total:"
            ws[f'B{row}'] = f"${(arqueos_report.get('total_difference_cents', 0) / 100):.2f}"
            row += 1
            
            ws[f'A{row}'] = "Diferencia Promedio:"
            ws[f'B{row}'] = f"${(arqueos_report.get('average_difference_cents', 0) / 100):.2f}"
            row += 1
            
            ws[f'A{row}'] = "Matches Perfectos:"
            ws[f'B{row}'] = arqueos_report.get('perfect_matches', 0)
            row += 1
            
            ws[f'A{row}'] = "Discrepancias:"
            ws[f'B{row}'] = arqueos_report.get('discrepancies', 0)
            row += 1
            
            ws[f'A{row}'] = "Tasa de Discrepancia:"
            ws[f'B{row}'] = f"{arqueos_report.get('discrepancy_rate', 0):.2f}%"
            row += 2
            
            # Recent arqueos table
            recent_arqueos = arqueos_report.get('recent_arqueos', [])
            if recent_arqueos:
                ws[f'A{row}'] = "ARQUEOS RECIENTES (Últimos 10)"
                ws[f'A{row}'].font = Font(bold=True, size=12)
                row += 1
                
                ws[f'A{row}'] = "Fecha"
                ws[f'B{row}'] = "Sistema"
                ws[f'C{row}'] = "Físico"
                ws[f'D{row}'] = "Diferencia"
                ws[f'E{row}'] = "Sucursal"
                for cell in ws[f'A{row}:E{row}']:
                    for c in cell:
                        c.fill = header_fill
                        c.font = header_font
                row += 1
                
                for arqueo in recent_arqueos:
                    date_str = arqueo.get('date', '')
                    if date_str:
                        try:
                            dt = datetime.fromisoformat(date_str.replace('Z', '+00:00'))
                            date_str = dt.strftime('%d/%m/%Y')
                        except:
                            pass
                    
                    ws[f'A{row}'] = date_str
                    ws[f'B{row}'] = f"${(arqueo.get('system_total_cents', 0) / 100):.2f}"
                    ws[f'C{row}'] = f"${(arqueo.get('physical_count_cents', 0) / 100):.2f}"
                    ws[f'D{row}'] = f"${(arqueo.get('difference_cents', 0) / 100):.2f}"
                    ws[f'E{row}'] = str(arqueo.get('sucursal_id', ''))[:8] + '...' if len(str(arqueo.get('sucursal_id', ''))) > 8 else str(arqueo.get('sucursal_id', ''))
                    row += 1
                row += 1
            
            # Breakdown by sucursal (if available)
            by_sucursal = arqueos_report.get('by_sucursal', {})
            if by_sucursal:
                ws[f'A{row}'] = "DESGLOSE POR SUCURSAL"
                ws[f'A{row}'].font = Font(bold=True, size=12)
                row += 1
                
                ws[f'A{row}'] = "Sucursal"
                ws[f'B{row}'] = "Arqueos"
                ws[f'C{row}'] = "Matches Perfectos"
                ws[f'D{row}'] = "Diferencia Total"
                for cell in ws[f'A{row}:D{row}']:
                    for c in cell:
                        c.fill = header_fill
                        c.font = header_font
                row += 1
                
                for sucursal_id, data in by_sucursal.items():
                    ws[f'A{row}'] = str(sucursal_id)[:8] + '...' if len(str(sucursal_id)) > 8 else str(sucursal_id)
                    ws[f'B{row}'] = data.get('count', 0)
                    ws[f'C{row}'] = data.get('perfect_matches', 0)
                    ws[f'D{row}'] = f"${(data.get('total_difference_cents', 0) / 100):.2f}"
                    row += 1
        
        # Auto-adjust columns
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_forecasting_sheet(self, wb: Workbook, forecasting_data: Dict[str, Any]) -> None:
        """Create forecasting predictions sheet in Excel."""
        ws = wb.create_sheet("Forecasting")
        
        # Header style
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Title
        ws['A1'] = f"Pronósticos Avanzados - {forecasting_data.get('forecast_days', 7)} días"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:F1')
        
        row = 3
        
        predictions = forecasting_data.get('predictions', {})
        forecast_days = forecasting_data.get('forecast_days', 7)
        
        # Process each module
        for module_name, module_predictions in predictions.items():
            if not module_predictions:
                continue
            
            # Module header
            ws[f'A{row}'] = f"MÓDULO: {module_name.upper()}"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            ws.merge_cells(f'A{row}:F{row}')
            row += 1
            
            # Sales predictions
            sales_pred = module_predictions.get('sales')
            if sales_pred and not sales_pred.get('error'):
                forecast = sales_pred.get('forecast', [])
                confidence = sales_pred.get('confidence', 'N/A')
                method = sales_pred.get('method', 'N/A')
                
                ws[f'A{row}'] = "Predicciones de Ventas"
                ws[f'A{row}'].font = Font(bold=True)
                row += 1
                
                ws[f'A{row}'] = "Confianza:"
                ws[f'B{row}'] = confidence
                ws[f'C{row}'] = "Método:"
                ws[f'D{row}'] = method
                row += 1
                
                ws[f'A{row}'] = "Fecha"
                ws[f'B{row}'] = "Revenue Previsto"
                ws[f'C{row}'] = "Ventas Previstas"
                ws[f'D{row}'] = "Factor Día"
                for cell in ws[f'A{row}:D{row}']:
                    for c in cell:
                        c.fill = header_fill
                        c.font = header_font
                row += 1
                
                for day in forecast[:forecast_days]:
                    date_str = day.get('date', '')
                    if date_str:
                        try:
                            dt = datetime.fromisoformat(date_str.replace('Z', '+00:00'))
                            date_str = dt.strftime('%d/%m/%Y')
                        except:
                            pass
                    
                    ws[f'A{row}'] = date_str
                    ws[f'B{row}'] = f"${((day.get('predicted_revenue_cents', 0) or 0) / 100):.2f}"
                    ws[f'C{row}'] = day.get('predicted_count', 0) or 0
                    ws[f'D{row}'] = f"{day.get('day_of_week_factor', 1.0):.2f}"
                    row += 1
                row += 1
            
            # Capacity predictions (only for recepcion or total)
            capacity_pred = module_predictions.get('capacity')
            if capacity_pred and not capacity_pred.get('error'):
                forecast = capacity_pred.get('forecast', [])
                confidence = capacity_pred.get('confidence', 'N/A')
                
                ws[f'A{row}'] = "Predicciones de Capacidad"
                ws[f'A{row}'].font = Font(bold=True)
                row += 1
                
                ws[f'A{row}'] = "Confianza:"
                ws[f'B{row}'] = confidence
                row += 1
                
                ws[f'A{row}'] = "Fecha"
                ws[f'B{row}'] = "Capacidad Prevista"
                ws[f'C{row}'] = "Utilización"
                for cell in ws[f'A{row}:C{row}']:
                    for c in cell:
                        c.fill = header_fill
                        c.font = header_font
                row += 1
                
                for day in forecast[:forecast_days]:
                    date_str = day.get('date', '')
                    if date_str:
                        try:
                            dt = datetime.fromisoformat(date_str.replace('Z', '+00:00'))
                            date_str = dt.strftime('%d/%m/%Y')
                        except:
                            pass
                    
                    ws[f'A{row}'] = date_str
                    ws[f'B{row}'] = day.get('predicted_capacity', 0) or 0
                    utilization = day.get('predicted_utilization_percent', 0) or 0
                    ws[f'C{row}'] = f"{utilization:.1f}%"
                    row += 1
                row += 1
            
            # Stock predictions (only for kidibar or total)
            stock_pred = module_predictions.get('stock')
            if stock_pred and not stock_pred.get('error'):
                suggestions = stock_pred.get('reorder_suggestions', [])
                confidence = stock_pred.get('confidence', 'N/A')
                
                ws[f'A{row}'] = "Sugerencias de Reorden"
                ws[f'A{row}'].font = Font(bold=True)
                row += 1
                
                ws[f'A{row}'] = "Confianza:"
                ws[f'B{row}'] = confidence
                row += 1
                
                if suggestions:
                    ws[f'A{row}'] = "Producto"
                    ws[f'B{row}'] = "Stock Actual"
                    ws[f'C{row}'] = "Cantidad Sugerida"
                    ws[f'D{row}'] = "Prioridad"
                    for cell in ws[f'A{row}:D{row}']:
                        for c in cell:
                            c.fill = header_fill
                            c.font = header_font
                    row += 1
                    
                    for suggestion in suggestions[:20]:  # Limit to top 20
                        product_name = suggestion.get('product_name', 'N/A')
                        ws[f'A{row}'] = product_name[:50]  # Truncate long names
                        ws[f'B{row}'] = suggestion.get('current_stock', 0) or 0
                        ws[f'C{row}'] = suggestion.get('suggested_quantity', 0) or 0
                        ws[f'D{row}'] = suggestion.get('priority', 'medium')
                        row += 1
                row += 1
            
            row += 1  # Space between modules
        
        # Auto-adjust columns
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # ========== PDF SECTION CREATION METHODS ==========
    
    def _add_sales_section_pdf(self, story: list, sales_report: Dict[str, Any], heading_style: ParagraphStyle) -> None:
        """Add sales section to PDF."""
        story.append(Paragraph("Ventas", heading_style))
        
        if sales_report:
            data = [
                ["Métrica", "Valor"],
                ["Total Revenue", f"${(sales_report.get('total_revenue_cents', 0) / 100):.2f}"],
                ["Total Ventas", str(sales_report.get('sales_count', 0))],
                ["Ticket Promedio", f"${(sales_report.get('average_transaction_value_cents', 0) / 100):.2f}"]
            ]
            
            table = Table(data, colWidths=[3*inch, 2*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(table)
            story.append(Spacer(1, 0.2 * inch))
    
    def _add_stock_section_pdf(self, story: list, stock_report: Dict[str, Any], heading_style: ParagraphStyle) -> None:
        """Add stock section to PDF."""
        story.append(Paragraph("Inventario", heading_style))
        
        if stock_report:
            data = [
                ["Métrica", "Valor"],
                ["Total Productos", str(stock_report.get('total_products', 0))],
                ["Valor Total", f"${(stock_report.get('total_stock_value_cents', 0) / 100):.2f}"],
                ["Alertas", str(stock_report.get('alerts_count', 0))]
            ]
            
            table = Table(data, colWidths=[3*inch, 2*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(table)
            story.append(Spacer(1, 0.2 * inch))
    
    def _add_services_section_pdf(self, story: list, services_report: Dict[str, Any], heading_style: ParagraphStyle) -> None:
        """Add services section to PDF."""
        story.append(Paragraph("Servicios", heading_style))
        
        if services_report:
            data = [
                ["Métrica", "Valor"],
                ["Timers Activos", str(services_report.get('active_timers_count', 0))],
                ["Total Servicios", str(services_report.get('total_services', 0))]
            ]
            
            table = Table(data, colWidths=[3*inch, 2*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(table)
            story.append(Spacer(1, 0.2 * inch))
    
    def _add_peak_hours_section_pdf(self, story: list, peak_hours_report: Dict[str, Any], heading_style: ParagraphStyle) -> None:
        """Add peak hours section to PDF."""
        story.append(Paragraph("Horas Pico", heading_style))
        
        if peak_hours_report:
            busiest_hour = peak_hours_report.get('busiest_hour', {})
            data = [
                ["Hora", "Ventas", "Revenue"],
                [f"{busiest_hour.get('hour', 0)}:00h", str(busiest_hour.get('sales_count', 0)), 
                 f"${(busiest_hour.get('revenue_cents', 0) / 100):.2f}"]
            ]
            
            # Add top hours
            if peak_hours_report.get('peak_hours'):
                for peak in peak_hours_report['peak_hours'][:5]:
                    data.append([
                        f"{peak.get('hour', 0)}:00h",
                        str(peak.get('sales_count', 0)),
                        f"${(peak.get('revenue_cents', 0) / 100):.2f}"
                    ])
            
            table = Table(data, colWidths=[2*inch, 1.5*inch, 1.5*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(table)
            story.append(Spacer(1, 0.2 * inch))
    
    def _add_top_products_section_pdf(self, story: list, top_products_report: Dict[str, Any], heading_style: ParagraphStyle) -> None:
        """Add top products section to PDF."""
        story.append(Paragraph(f"Productos Top - Últimos {top_products_report.get('period_days', 7)} días", heading_style))
        
        if top_products_report and top_products_report.get('top_products'):
            data = [["Rank", "Producto", "Cantidad", "Revenue"]]
            
            for idx, product in enumerate(top_products_report['top_products'], 1):
                data.append([
                    str(idx),
                    product.get('product_name', ''),
                    str(product.get('quantity_sold', 0)),
                    f"${(product.get('revenue_cents', 0) / 100):.2f}"
                ])
            
            table = Table(data, colWidths=[0.5*inch, 2.5*inch, 1*inch, 1*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(table)
            story.append(Spacer(1, 0.2 * inch))
    
    def _add_top_services_section_pdf(self, story: list, top_services_report: Dict[str, Any], heading_style: ParagraphStyle) -> None:
        """Add top services section to PDF."""
        story.append(Paragraph(f"Servicios Top - Últimos {top_services_report.get('period_days', 7)} días", heading_style))
        
        if top_services_report and top_services_report.get('top_services'):
            data = [["Rank", "Servicio", "Usos", "Duración Promedio"]]
            
            for idx, service in enumerate(top_services_report['top_services'], 1):
                data.append([
                    str(idx),
                    service.get('service_name', ''),
                    str(service.get('usage_count', 0)),
                    f"{service.get('avg_duration_minutes', 0):.1f} min"
                ])
            
            table = Table(data, colWidths=[0.5*inch, 2.5*inch, 1*inch, 1*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(table)
            story.append(Spacer(1, 0.2 * inch))
    
    def _add_top_customers_section_pdf(self, story: list, top_customers_report: Dict[str, Any], heading_style: ParagraphStyle) -> None:
        """Add top customers section to PDF."""
        story.append(Paragraph(f"Clientes Top - Últimos {top_customers_report.get('period_days', 7)} días", heading_style))
        
        if top_customers_report and top_customers_report.get('top_customers'):
            data = [["Rank", "Nombre", "Edad", "Visitas", "Total Gastado"]]
            
            for idx, customer in enumerate(top_customers_report['top_customers'], 1):
                data.append([
                    str(idx),
                    customer.get('child_name', ''),
                    str(customer.get('child_age', '')) if customer.get('child_age') else 'N/A',
                    str(customer.get('visit_count', 0)),
                    f"${((customer.get('total_revenue_cents', 0) or 0) / 100):.2f}"
                ])
            
            table = Table(data, colWidths=[0.5*inch, 2*inch, 0.8*inch, 0.8*inch, 1*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(table)
            story.append(Spacer(1, 0.2 * inch))
    
    def _add_executive_summary_section_pdf(self, story: list, summary_report: Dict[str, Any], heading_style: ParagraphStyle) -> None:
        """Add executive summary section to PDF."""
        story.append(Paragraph("Resumen Ejecutivo", heading_style))
        
        if not summary_report:
            from reportlab.lib.styles import getSampleStyleSheet
            styles = getSampleStyleSheet()
            story.append(Paragraph("No hay datos disponibles para el resumen ejecutivo.", styles['Normal']))
            story.append(Spacer(1, 0.2 * inch))
            return
        
        from reportlab.lib.styles import getSampleStyleSheet
        styles = getSampleStyleSheet()
        
        # Format generated_at date
        generated_at_str = "N/A"
        if summary_report.get('generated_at'):
            try:
                dt = datetime.fromisoformat(summary_report['generated_at'].replace('Z', '+00:00'))
                generated_at_str = dt.strftime('%d/%m/%Y %H:%M:%S')
            except:
                generated_at_str = str(summary_report.get('generated_at', 'N/A'))
        
        # Generated at info
        story.append(Paragraph(f"<b>Generado:</b> {generated_at_str}", styles['Normal']))
        story.append(Spacer(1, 0.2 * inch))
        
        # Sales section
        sales = summary_report.get('sales')
        if sales:
            story.append(Paragraph("<b>Ventas</b>", styles['Heading3']))
            data = [
                ["Métrica", "Valor"],
                ["Total Revenue", f"${(sales.get('total_revenue_cents', 0) / 100):.2f}"],
                ["Total Ventas", str(sales.get('sales_count', 0))],
                ["Ticket Promedio", f"${(sales.get('average_transaction_value_cents', 0) / 100):.2f}"]
            ]
            
            table = Table(data, colWidths=[3*inch, 2*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(table)
            story.append(Spacer(1, 0.15 * inch))
        
        # Stock section
        stock = summary_report.get('stock')
        if stock:
            story.append(Paragraph("<b>Inventario</b>", styles['Heading3']))
            data = [
                ["Métrica", "Valor"],
                ["Total Productos", str(stock.get('total_products', 0))],
                ["Valor Total", f"${(stock.get('total_stock_value_cents', 0) / 100):.2f}"],
                ["Alertas", str(stock.get('alerts_count', 0))]
            ]
            
            table = Table(data, colWidths=[3*inch, 2*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(table)
            story.append(Spacer(1, 0.15 * inch))
        
        # Services section
        services = summary_report.get('services')
        if services:
            story.append(Paragraph("<b>Servicios</b>", styles['Heading3']))
            data = [
                ["Métrica", "Valor"],
                ["Timers Activos", str(services.get('active_timers_count', 0))],
                ["Total Servicios", str(services.get('total_services', 0))]
            ]
            
            table = Table(data, colWidths=[3*inch, 2*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(table)
            story.append(Spacer(1, 0.15 * inch))
        
        story.append(Spacer(1, 0.2 * inch))
    
    def _add_customers_summary_section_pdf(self, story: list, customers_summary: Dict[str, Any], heading_style: ParagraphStyle) -> None:
        """Add customers summary section to PDF."""
        story.append(Paragraph("Resumen de Clientes", heading_style))
        
        if not customers_summary:
            from reportlab.lib.styles import getSampleStyleSheet
            styles = getSampleStyleSheet()
            story.append(Paragraph("No hay datos disponibles para el resumen de clientes.", styles['Normal']))
            story.append(Spacer(1, 0.2 * inch))
            return
        
        from reportlab.lib.styles import getSampleStyleSheet
        styles = getSampleStyleSheet()
        
        # Main metrics
        avg_revenue = customers_summary.get('avg_revenue_per_customer_cents', 0) or 0
        total_revenue = customers_summary.get('total_revenue_cents', 0) or 0
        recepcion_count = customers_summary.get('recepcion_customers', 0) or 0
        kidibar_count = customers_summary.get('kidibar_customers', 0) or 0
        
        data = [
            ["Métrica", "Valor"],
            ["Total Clientes Únicos", str(customers_summary.get('total_unique_customers', 0))],
            ["Nuevos Clientes", str(customers_summary.get('new_customers', 0))],
            ["Revenue Promedio por Cliente", f"${(avg_revenue / 100):.2f}"],
            ["Revenue Total", f"${(total_revenue / 100):.2f}"]
        ]
        
        table = Table(data, colWidths=[3*inch, 2*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(table)
        story.append(Spacer(1, 0.2 * inch))
        
        # Breakdown by module
        if recepcion_count > 0 or kidibar_count > 0:
            story.append(Paragraph("<b>Desglose por Módulo</b>", styles['Heading3']))
            module_data = [
                ["Módulo", "Clientes"],
                ["Recepción", str(recepcion_count)],
                ["KidiBar", str(kidibar_count)]
            ]
            
            module_table = Table(module_data, colWidths=[3*inch, 2*inch])
            module_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(module_table)
            story.append(Spacer(1, 0.2 * inch))
    
    def _add_arqueos_section_pdf(self, story: list, arqueos_report: Dict[str, Any], heading_style: ParagraphStyle) -> None:
        """Add arqueos section to PDF."""
        story.append(Paragraph("Arqueos", heading_style))
        
        if not arqueos_report:
            from reportlab.lib.styles import getSampleStyleSheet
            styles = getSampleStyleSheet()
            story.append(Paragraph("No hay datos disponibles para arqueos.", styles['Normal']))
            story.append(Spacer(1, 0.2 * inch))
            return
        
        from reportlab.lib.styles import getSampleStyleSheet
        styles = getSampleStyleSheet()
        
        # Period info
        period = arqueos_report.get('period', {})
        if period:
            start_date_str = period.get('start_date', 'N/A')
            end_date_str = period.get('end_date', 'N/A')
            story.append(Paragraph(f"<b>Período:</b> {start_date_str} - {end_date_str}", styles['Normal']))
            story.append(Spacer(1, 0.1 * inch))
        
        # Summary metrics
        summary_data = [
            ["Métrica", "Valor"],
            ["Total Arqueos", str(arqueos_report.get('total_arqueos', 0))],
            ["Total Sistema", f"${(arqueos_report.get('total_system_cents', 0) / 100):.2f}"],
            ["Total Físico", f"${(arqueos_report.get('total_physical_cents', 0) / 100):.2f}"],
            ["Diferencia Total", f"${(arqueos_report.get('total_difference_cents', 0) / 100):.2f}"],
            ["Diferencia Promedio", f"${(arqueos_report.get('average_difference_cents', 0) / 100):.2f}"],
            ["Matches Perfectos", str(arqueos_report.get('perfect_matches', 0))],
            ["Discrepancias", str(arqueos_report.get('discrepancies', 0))],
            ["Tasa de Discrepancia", f"{arqueos_report.get('discrepancy_rate', 0):.2f}%"]
        ]
        
        summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 0.2 * inch))
        
        # Recent arqueos table
        recent_arqueos = arqueos_report.get('recent_arqueos', [])
        if recent_arqueos:
            story.append(Paragraph("<b>Arqueos Recientes (Últimos 10)</b>", styles['Heading3']))
            recent_data = [["Fecha", "Sistema", "Físico", "Diferencia"]]
            
            for arqueo in recent_arqueos:
                date_str = arqueo.get('date', '')
                if date_str:
                    try:
                        dt = datetime.fromisoformat(date_str.replace('Z', '+00:00'))
                        date_str = dt.strftime('%d/%m/%Y')
                    except:
                        pass
                
                recent_data.append([
                    date_str,
                    f"${(arqueo.get('system_total_cents', 0) / 100):.2f}",
                    f"${(arqueo.get('physical_count_cents', 0) / 100):.2f}",
                    f"${(arqueo.get('difference_cents', 0) / 100):.2f}"
                ])
            
            recent_table = Table(recent_data, colWidths=[1.5*inch, 1.5*inch, 1.5*inch, 1.5*inch])
            recent_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(recent_table)
            story.append(Spacer(1, 0.2 * inch))
        
        # Breakdown by sucursal (if available)
        by_sucursal = arqueos_report.get('by_sucursal', {})
        if by_sucursal:
            story.append(Paragraph("<b>Desglose por Sucursal</b>", styles['Heading3']))
            sucursal_data = [["Sucursal", "Arqueos", "Matches Perfectos", "Diferencia Total"]]
            
            for sucursal_id, data in by_sucursal.items():
                sucursal_str = str(sucursal_id)[:8] + '...' if len(str(sucursal_id)) > 8 else str(sucursal_id)
                sucursal_data.append([
                    sucursal_str,
                    str(data.get('count', 0)),
                    str(data.get('perfect_matches', 0)),
                    f"${(data.get('total_difference_cents', 0) / 100):.2f}"
                ])
            
            sucursal_table = Table(sucursal_data, colWidths=[2*inch, 1*inch, 1.5*inch, 1.5*inch])
            sucursal_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(sucursal_table)
            story.append(Spacer(1, 0.2 * inch))
    
    def _add_forecasting_section_pdf(self, story: list, forecasting_data: Dict[str, Any], heading_style: ParagraphStyle) -> None:
        """Add forecasting section to PDF."""
        forecast_days = forecasting_data.get('forecast_days', 7)
        story.append(Paragraph(f"Forecasting - Pronósticos {forecast_days} días", heading_style))
        
        predictions = forecasting_data.get('predictions', {})
        if not predictions:
            from reportlab.lib.styles import getSampleStyleSheet
            styles = getSampleStyleSheet()
            story.append(Paragraph("No hay predicciones disponibles.", styles['Normal']))
            story.append(Spacer(1, 0.2 * inch))
            return
        
        from reportlab.lib.styles import getSampleStyleSheet
        styles = getSampleStyleSheet()
        
        # Create style for subheadings
        subheading_style = ParagraphStyle(
            'CustomSubHeading',
            parent=styles['Heading3'],
            fontSize=12,
            textColor=colors.HexColor('#16213e'),
            spaceAfter=6,
            spaceBefore=8
        )
        
        # Process each module
        for module_name, module_predictions in predictions.items():
            if not module_predictions:
                continue
            
            # Module header
            story.append(Paragraph(f"<b>Módulo: {module_name.upper()}</b>", subheading_style))
            
            # Sales predictions
            sales_pred = module_predictions.get('sales')
            if sales_pred and not sales_pred.get('error'):
                forecast = sales_pred.get('forecast', [])
                confidence = sales_pred.get('confidence', 'N/A')
                method = sales_pred.get('method', 'N/A')
                
                story.append(Paragraph(f"Predicciones de Ventas (Confianza: {confidence}, Método: {method})", styles['Normal']))
                
                if forecast:
                    data = [["Fecha", "Revenue Previsto", "Ventas Previstas"]]
                    for day in forecast[:forecast_days]:
                        date_str = day.get('date', '')
                        if date_str:
                            try:
                                dt = datetime.fromisoformat(date_str.replace('Z', '+00:00'))
                                date_str = dt.strftime('%d/%m/%Y')
                            except:
                                pass
                        
                        data.append([
                            date_str,
                            f"${((day.get('predicted_revenue_cents', 0) or 0) / 100):.2f}",
                            str(day.get('predicted_count', 0) or 0)
                        ])
                    
                    if len(data) > 1:
                        table = Table(data, colWidths=[2*inch, 1.5*inch, 1.5*inch])
                        table.setStyle(TableStyle([
                            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                            ('FONTSIZE', (0, 0), (-1, 0), 11),
                            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                            ('GRID', (0, 0), (-1, -1), 1, colors.black)
                        ]))
                        story.append(table)
                        story.append(Spacer(1, 0.15 * inch))
            
            # Capacity predictions (only for recepcion or total)
            capacity_pred = module_predictions.get('capacity')
            if capacity_pred and not capacity_pred.get('error'):
                forecast = capacity_pred.get('forecast', [])
                confidence = capacity_pred.get('confidence', 'N/A')
                
                story.append(Paragraph(f"Predicciones de Capacidad (Confianza: {confidence})", styles['Normal']))
                
                if forecast:
                    data = [["Fecha", "Capacidad Prevista", "Utilización"]]
                    for day in forecast[:forecast_days]:
                        date_str = day.get('date', '')
                        if date_str:
                            try:
                                dt = datetime.fromisoformat(date_str.replace('Z', '+00:00'))
                                date_str = dt.strftime('%d/%m/%Y')
                            except:
                                pass
                        
                        utilization = day.get('predicted_utilization_percent', 0) or 0
                        data.append([
                            date_str,
                            str(day.get('predicted_capacity', 0) or 0),
                            f"{utilization:.1f}%"
                        ])
                    
                    if len(data) > 1:
                        table = Table(data, colWidths=[2*inch, 1.5*inch, 1.5*inch])
                        table.setStyle(TableStyle([
                            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                            ('FONTSIZE', (0, 0), (-1, 0), 11),
                            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                            ('GRID', (0, 0), (-1, -1), 1, colors.black)
                        ]))
                        story.append(table)
                        story.append(Spacer(1, 0.15 * inch))
            
            # Stock predictions (only for kidibar or total)
            stock_pred = module_predictions.get('stock')
            if stock_pred and not stock_pred.get('error'):
                suggestions = stock_pred.get('reorder_suggestions', [])
                confidence = stock_pred.get('confidence', 'N/A')
                
                story.append(Paragraph(f"Sugerencias de Reorden (Confianza: {confidence})", styles['Normal']))
                
                if suggestions:
                    data = [["Producto", "Stock Actual", "Cantidad Sugerida", "Prioridad"]]
                    for suggestion in suggestions[:15]:  # Limit to top 15 for PDF
                        product_name = suggestion.get('product_name', 'N/A')
                        data.append([
                            product_name[:30],  # Truncate long names
                            str(suggestion.get('current_stock', 0) or 0),
                            str(suggestion.get('suggested_quantity', 0) or 0),
                            suggestion.get('priority', 'medium')
                        ])
                    
                    if len(data) > 1:
                        table = Table(data, colWidths=[2*inch, 1*inch, 1*inch, 1*inch])
                        table.setStyle(TableStyle([
                            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                            ('FONTSIZE', (0, 0), (-1, 0), 10),
                            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                            ('GRID', (0, 0), (-1, -1), 1, colors.black)
                        ]))
                        story.append(table)
                        story.append(Spacer(1, 0.15 * inch))
            
            story.append(Spacer(1, 0.2 * inch))  # Space between modules
        
        story.append(Spacer(1, 0.2 * inch))
    
    def _add_predictions_section_pdf(self, story: list, predictions: Dict[str, Any], heading_style: ParagraphStyle) -> None:
        """Add predictions section to PDF."""
        from reportlab.lib.styles import getSampleStyleSheet
        styles = getSampleStyleSheet()
        
        story.append(Paragraph("Predicciones y Análisis", heading_style))
        story.append(Spacer(1, 0.1 * inch))
        
        if not predictions:
            story.append(Paragraph("No hay predicciones disponibles.", styles['Normal']))
            story.append(Spacer(1, 0.2 * inch))
            return
        
        # Create style for subheadings
        subheading_style = ParagraphStyle(
            'CustomSubHeading',
            parent=styles['Heading3'],
            fontSize=12,
            textColor=colors.HexColor('#16213e'),
            spaceAfter=6,
            spaceBefore=8
        )
        
        # Sales predictions
        if predictions.get('sales') and predictions['sales'].get('forecast'):
            story.append(Paragraph("Predicciones de Ventas", subheading_style))
            forecast = predictions['sales']['forecast']
            confidence = predictions['sales'].get('confidence', 'N/A')
            
            data = [["Fecha", "Revenue Previsto", "Ventas Previstas"]]
            for day in forecast[:7]:  # Show first 7 days
                date_str = day.get('date', '')
                if date_str:
                    try:
                        # Parse ISO date and format as DD/MM/YYYY
                        from datetime import datetime
                        dt = datetime.fromisoformat(date_str.replace('Z', '+00:00'))
                        date_str = dt.strftime('%d/%m/%Y')
                    except:
                        pass
                
                revenue_cents = day.get('predicted_revenue_cents', 0) or 0
                count = day.get('predicted_count', 0) or 0
                data.append([
                    date_str,
                    f"${(revenue_cents / 100):.2f}",
                    str(count)
                ])
            
            if len(data) > 1:
                table = Table(data, colWidths=[2*inch, 1.5*inch, 1.5*inch])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 11),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                story.append(table)
                story.append(Paragraph(f"Confianza: {confidence}", styles['Normal']))
                story.append(Spacer(1, 0.2 * inch))
        
        # Peak hours predictions
        if predictions.get('peak_hours') and predictions['peak_hours'].get('forecast'):
            story.append(Paragraph("Predicciones de Horas Pico", subheading_style))
            forecast = predictions['peak_hours']['forecast']
            
            data = [["Hora", "Ventas Previstas"]]
            for hour_pred in forecast[:10]:  # Show top 10 hours
                hour = hour_pred.get('hour', 0)
                count = hour_pred.get('predicted_count', 0) or 0
                data.append([
                    f"{hour}:00h",
                    str(count)
                ])
            
            if len(data) > 1:
                table = Table(data, colWidths=[2*inch, 2*inch])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 11),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                story.append(table)
                story.append(Spacer(1, 0.2 * inch))
        
        # Stock predictions (reorder suggestions)
        if predictions.get('stock') and predictions['stock'].get('reorder_suggestions'):
            story.append(Paragraph("Sugerencias de Reorden", subheading_style))
            suggestions = predictions['stock']['reorder_suggestions']
            
            data = [["Producto", "Stock Actual", "Cantidad Sugerida"]]
            for suggestion in suggestions[:10]:  # Show top 10 suggestions
                product_name = suggestion.get('product_name', 'N/A')
                current_stock = suggestion.get('current_stock', 0) or 0
                suggested_qty = suggestion.get('suggested_quantity', 0) or 0
                data.append([
                    product_name[:30],  # Truncate long names
                    str(current_stock),
                    str(suggested_qty)
                ])
            
            if len(data) > 1:
                table = Table(data, colWidths=[2.5*inch, 1.5*inch, 1.5*inch])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 11),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                story.append(table)
                story.append(Spacer(1, 0.2 * inch))
